import re
from typing import Tuple, List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from matplotlib import pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit
import os.path as pth


class Report:
    """Класс, формирующий отчеты из данных по вакансиям"""

    def __init__(self, salary_by_year: Dict[str, int],
                 count_by_year: Dict[str, int], prof_salary_by_year: Dict[str, int],
                 prof_count_by_year: Dict[str, int], prof_name: str):
        """Инициализирует объект Report

            Args:
                salary_by_city (Dict[str, int]): З\п по городам
                count_by_city (Dict[str, int]): Количество вакансий по гродам
                salary_by_year (Dict[str, int]): З\п по годам
                count_by_year (Dict[str, int]): Количество вакансий по годам
                prof_salary_by_year (Dict[str, int]): З\п по професси по годам
                prof_count_by_year (Dict[str, int]): Количество вакансий по профессии по годам
                prof_name (str): Профессия
        """
        self.__salary_by_year = salary_by_year
        self.__count_by_year = count_by_year
        self.__prof_salary_by_year = prof_salary_by_year
        self.__prof_count_by_year = prof_count_by_year
        self.__prof_name = prof_name
        self.__font = Font(bold=True)
        self.__border = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.__fig, self.__axs = plt.subplots(2, 2, layout='tight', figsize=[10, 10])

    def generate_excel(self, years_table_header: List[Tuple], years_table: List[Tuple]):
        """Создает excel файл с таблицами по пути report/report.xlsx"""
        wk = Workbook()
        dest_filename = pth.relpath(pth.join('report', 'report.xlsx'))
        ws = wk.worksheets[0]
        ws.title = "Статистика по годам"
        ws.append(years_table_header)
        self.__style_header(ws)

        for t in years_table:
            ws.append(t)
        self.__style_cells(ws)

        if check_file('xlsx', dest_filename):
            wk.save(filename=dest_filename)

    def generate_image(self, salary_by_year: Dict[str, int], prof_salary_by_year: Dict[str, int],
                       count_by_year: Dict[str, int], prof_count_by_year: Dict[str, int]) -> None:
        """Создает изображение графиков по пути report/graph.png"""
        self.__generate_salary_diagram(salary_by_year, prof_salary_by_year)
        self.__generate_vacancy_diagram(count_by_year, prof_count_by_year)
        save_path = pth.relpath(pth.join('report', 'graph.png'))
        if check_file('png', save_path):
            plt.savefig(save_path)

    def generate_pdf(self, wkhtml_path: str, years_table_header, years_table) -> None:
        """Создает pdf файл по пути report/report.pdf с помощью таблиц и шаблона по пути template/template.html

            Args:
                wkhtml_path (str): Путь к модулю wkghtml
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template('template/template_lite.html')

        pdf_template = template.render({
            'first_table': years_table,
            'first_table_header': years_table_header,
            'prof_name': self.__prof_name,
            'to_css': pth.abspath(pth.join('template', 'style.css')),
            'to_img': pth.abspath(pth.join('report', 'graph.png'))
        })
        config = pdfkit.configuration(wkhtmltopdf=wkhtml_path)
        save_path = pth.normpath(pth.join('report', 'report.pdf'))
        if check_file('pdf', save_path):
            pdfkit.from_string(pdf_template, save_path, configuration=config,
                               options={"enable-local-file-access": ""})

    def generate_years_table(self) -> (Tuple[str, str, str, str, str], List[Tuple[int, int, int, int, int]]):
        """Создает заголовок и генератор строк для количества вакансий и з\п по вакансиям и профессии по годам.

                    Returns:
                        Tuple[
                            Tuple[str, str, str, str, str]: Заголовок,
                            List[
                                Tuple[
                                    int: Год,
                                    int: З\п по годам,
                                    int: З\п по годам для профессии,
                                    int: Количество вакансий по годам,
                                    int: Количество вакансий по годам для профессии
                                ]: Строка таблицы
                            ]
                        ]
                """
        fth = ('Год', 'Средняя зарплата', f'Средняя зарплата - {self.__prof_name}', 'Количество вакансий',
               f'Количество вакансий - {self.__prof_name}')
        years = {k: k for k in self.__count_by_year.keys()}
        ft = [(years[i], self.__salary_by_year[i], self.__prof_salary_by_year[i],
               self.__count_by_year[i], self.__prof_count_by_year[i]) for i in years]
        return fth, ft

    def __style_header(self, worksheet: Worksheet) -> None:
        """Применяет стили шрифта и рамки для главной строки таблицы

            Args:
                worksheet (Worksheet): Страница excell
        """
        for header in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in header:
                cell.font = self.__font
                cell.border = self.__border

    def __style_cells(self, worksheet: Worksheet) -> None:
        """Применяет рамку и устанавливает ширину по максимално широкой ячейке для каждого столбца

            Args:
                worksheet (Worksheet): Страница excell
        """
        dims = []
        for column in worksheet.columns:
            max_width = 0
            for cell in column:
                if cell.value != " ":
                    cell.border = self.__border
                    if max_width < len(str(cell.value)):
                        max_width = len(str(cell.value))
            dims.append(max_width + 2)
        for i, width in enumerate(dims):
            worksheet.column_dimensions[get_column_letter(i + 1)].width = width

    def __style_percent_column(self, worksheet: Worksheet, col_index: int) -> None:
        """Применяет ко всем ячейкам столбца стиль процентный с двумя знаками после запятой

            Args:
                worksheet (Worksheet): Страница excell
                col_index (int): Индекс столбца
        """
        for column in worksheet.iter_cols(col_index):
            for cell in column:
                cell.style = 'Percent'
                cell.number_format = '0.00%'

    def __generate_salary_diagram(self, salary_by_year: Dict[str, int], prof_salary_by_year: Dict[str, int]):
        """Создает диаграмму з\п по годам в первом окне"""
        ax = self.__axs[0, 0]
        ax.set_title('Уровень зарплат по годам', fontsize=20)
        y_pos = range(len(salary_by_year.keys()))
        years = list(salary_by_year.keys())
        _salary_by_year = list(salary_by_year.values())
        prof_salary = list(prof_salary_by_year.values())
        overall = ax.bar(y_pos, _salary_by_year, width=0.3)
        prof = ax.bar(list(map(lambda x: x + 0.3, y_pos)), prof_salary, width=0.3)
        ax.set_xticks(list(map(lambda x: x + 0.15, y_pos)), labels=years, rotation="vertical", fontsize=8)
        for tick in ax.get_yticklabels():
            tick.set_fontsize(8)
        ax.legend([overall, prof], [r'средняя з\п', r'з\п ' + self.__prof_name], fontsize=8)
        ax.grid(axis='y')

    def __generate_vacancy_diagram(self, count_by_year: Dict[str, int], prof_count_by_year: Dict[str, int]):
        """Создает диаграмму количества вакансий по годам во втором окне"""
        ax = self.__axs[0, 1]
        y_pos = range(len(count_by_year.keys()))
        years = list(count_by_year.keys())
        salary_by_year = list(count_by_year.values())
        prof_salary = list(prof_count_by_year.values())
        overall = ax.bar(y_pos, salary_by_year, width=0.3)
        prof = ax.bar(list(map(lambda x: x + 0.3, y_pos)), prof_salary, width=0.3)
        ax.set_xticks(list(map(lambda x: x + 0.15, y_pos)), labels=years, rotation="vertical", fontsize=8)
        for tick in ax.get_yticklabels():
            tick.set_fontsize(8)
        ax.legend([overall, prof], ['Количество вакансий', 'Количество ваканисий\n' + self.__prof_name], fontsize=8)
        ax.set_title('Количество вакансий по годам', fontsize=20)
        ax.grid(axis='y')


def check_file(ext: str, dir_name: str) -> bool:
    """Проверяет, что названия файла имеет правильное расширение, в качестве пути к файлу передана строка,
        папка для файла существует и что файла с таким именем не существует

        Args:
            ext (str): Расширение файла
            dir_name (str): Путь, куда нужно сохранить файл

        Returns:
            bool: Да, если путь и расширение валидны
    """
    if not isinstance(dir_name, str):
        raise TypeError('')
    if not pth.basename(dir_name).endswith(f".{ext}"):
        raise KeyError('')
    if not pth.exists(pth.dirname(dir_name)):
        raise FileNotFoundError('Папки не существует')
    if pth.exists(dir_name):
        raise FileExistsError(f'{dir_name} уже существует')
    return True
